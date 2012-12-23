# -*- coding: utf-8  -*-
"""
Robot which will does substitutions of tags within wiki page content with external or
other wiki text data. Like dynamic text updating.

Look at http://de.wikipedia.org/wiki/Benutzer:DrTrigon/Benutzer:DrTrigonBot/config.css
for 'postproc' example code.

Look at https://wiki.toolserver.org/view/Mail how to setup mail handling. The
following code was used in file '$HOME/.forward+subster':
--- --- --- --- --- --- --- --- --- ---
> ~/data/subster/mail_inbox
--- --- --- --- --- --- --- --- --- ---
in order to enable mail (mbox style) storage in given file for address:
drtrigon+subster@toolserver.org

Other scripts and tools related to this bot are:
- subster_irc.py            IRC Robot
- substersim.py             Subster simulation panel
- subster_mail_queue.py     Subster mail queue
"""
## @package subster
#  @brief   Dynamic Text Substitutions Robot
#
#  @copyright Dr. Trigon, 2009-2012
#
#  @section FRAMEWORK
#
#  Python wikipedia robot framework, DrTrigonBot.
#  @see http://pywikipediabot.sourceforge.net/
#  @see http://de.wikipedia.org/wiki/Benutzer:DrTrigonBot
#
#  External code / other modules used are listed here.
#  @see https://bitbucket.org/ericgazoni/openpyxl/wiki/Home
#  @see http://pypi.python.org/pypi/crontab/.11
#
#  @section LICENSE
#
#  Distributed under the terms of the MIT license.
#  @see http://de.wikipedia.org/wiki/MIT-Lizenz
#
__version__ = '$Id$'
#


import re, sys, os, string, time
import difflib
import BeautifulSoup
import StringIO, zipfile, csv, urllib
import mailbox, mimetypes, datetime, email.utils
import openpyxl.reader.excel
import crontab
import logging
import copy
import ast

import pagegenerators, basic
# Splitting the bot into library parts
import wikipedia as pywikibot
from pywikibot import i18n
from pywikibot.comms import http


bot_config = {    # unicode values
        'TemplateName':     u'User:DrTrigonBot/Subster',    # or 'template' for 'Flagged Revisions'
        'ErrorTemplate':    u'----\n<b>SubsterBot Exception in "%s" (%s)</b>\n%s',

        # important to use a '.css' page here, since it HAS TO BE protected to
        # prevent malicious code injection !
        'ConfCSSpostproc':  u'User:DrTrigon/DrTrigonBot/subster-postproc.css',
        'ConfCSSconfig':    u'User:DrTrigon/DrTrigonBot/subster-config.css',

        'CodeTemplate':     u'\n%s(DATA, *args)\n',
        'CRONMaxDelay':     1*24*60*60,       # bot runs daily

        # regex values
        'var_regex_str':    u'<!--SUBSTER-%(var1)s-->%(cont)s<!--SUBSTER-%(var2)s-->',

        'mbox_file':        'mail_inbox',    # "drtrigon+subster@toolserver.org"
        'data_path':        '../data/subster',

        # bot paramater/options
        'param_default':    {
            'url':             '',
            'regex':           '',
            'value':           '',
            'count':           '0',
            #'postproc':        '("","")',
            'postproc':        '(\'\', \'\')',
            'wiki':            'False',         # may be change to url='wiki://'
            'magicwords_only': 'False',
            'beautifulsoup':   'False',         # DRTRIGON-88
            'expandtemplates': 'False',         # DRTRIGON-93 (only with 'wiki')
            'simple':          '',              # DRTRIGON-85
            'zip':             'False',
            'xlsx':            '',              #
            'ods':             '',              #
            # may be 'hours' have to be added too (e.g. for 'ar')
            'cron':            '',              # DRTRIGON-102
            'error':           repr('<noinclude>\n%(error)s\n</noinclude>'), # DRTRIGON-116
            #'djvu': ... u"djvused -e 'n' \"%s\"" ... djvutext.py
            #'pdf': ... u"pdftotext" or python module
            #'imageocr', 'swfocr', ...
        },

        # this is a system parameter and should not be changed! (copy.deepcopy)
        'EditFlags':        {'minorEdit': True, 'botflag': True},
}

## used/defined magic words, look also at bot_control
#  use, for example: '\<!--SUBSTER-BOTerror--\>\<!--SUBSTER-BOTerror--\>'
magic_words = {} # no magic word substitution (for empty dict)

# debug tools
# (look at 'bot_control.py' for more info)
debug = []
#debug.append( 'code' )  # code debugging


class SubsterBot(basic.AutoBasicBot):
    '''
    Robot which will does substitutions of tags within wiki page content with external or
    other wiki text data. Like dynamic text updating.
    '''

    _param_default = bot_config['param_default']

    _var_regex_str = bot_config['var_regex_str']%{'var1':'%(var)s','var2':'%(var)s','cont':'%(cont)s'}
    _BS_regex_str  = bot_config['var_regex_str']%{'var1':'%(var1)s','var2':'%(var2)sBS:/','cont':'%(cont)s'}

    # -template and subst-tag handling taken from MerlBot
    # -this bot could also be runned on my local wiki with an anacron-job

    def __init__(self):
        '''Constructor of SubsterBot(), initialize needed vars.'''

        pywikibot.output(u'\03{lightgreen}* Initialization of bot:\03{default}')

        logging.basicConfig(level=logging.DEBUG if ('code' in debug) else logging.INFO)

        basic.AutoBasicBot.__init__(self)

        # modification of timezone to be in sync with wiki
        os.environ['TZ'] = 'Europe/Amsterdam'
        time.tzset()
        pywikibot.output(u'Setting process TimeZone (TZ): %s' % str(time.tzname))    # ('CET', 'CEST')

        # init constants
        self._bot_config = bot_config
        # convert e.g. namespaces to corret language
        self._bot_config['TemplateName'] = pywikibot.Page(self.site, self._bot_config['TemplateName']).title()
        self._template_regex = re.compile('\{\{' + self._bot_config['TemplateName'] + '(.*?)\}\}', re.S)

        self._debug = debug

        # init constants
        self._userListPage        = pywikibot.Page(self.site, bot_config['TemplateName'])
        self._ConfCSSpostprocPage = pywikibot.Page(self.site, bot_config['ConfCSSpostproc'])
        self._ConfCSSconfigPage   = pywikibot.Page(self.site, bot_config['ConfCSSconfig'])
        self.pagegen     = pagegenerators.ReferringPageGenerator(self._userListPage, onlyTemplateInclusion=True)
        self._code       = self._ConfCSSpostprocPage.get()
        pywikibot.output(u'Imported postproc %s rev %s from %s' % \
          ((self._ConfCSSpostprocPage.title(asLink=True),) + self._ConfCSSpostprocPage.getVersionHistory(revCount=1)[0][:2]) )
        self._flagenable = {}
        if self._ConfCSSconfigPage.exists():
            exec(self._ConfCSSconfigPage.get())    # with variable: bot_config_wiki
            self._flagenable = bot_config_wiki['flagenable']
            pywikibot.output(u'Imported config %s rev %s from %s' % \
              ((self._ConfCSSconfigPage.title(asLink=True),) + self._ConfCSSconfigPage.getVersionHistory(revCount=1)[0][:2]) )

    def run(self, sim=False, msg=None, EditFlags=bot_config['EditFlags']):
        '''Run SubsterBot().'''

        pywikibot.output(u'\03{lightgreen}* Processing Template Backlink List:\03{default}')

        if sim:    self.pagegen = ['dummy']

        for page in self.pagegen:
            # setup source to get data from
            if sim:
                content = sim['content']
                params = [ sim ]
            else:
                pywikibot.output(u'Getting page "%s" via API from %s...'
                                 % (page.title(asLink=True), self.site))

                # get page content and operating mode
                content = self.load(page)
                params = self.loadTemplates(page, self._bot_config['TemplateName'],
                                            default=self._param_default)

            if not params: continue

            (substed_content, substed_tags) = self.subContent(content, params)

            # output result to page or return directly
            if sim:
                return substed_content
            elif (self.site.family.name == 'wikidata'):     # DRTRIGON-130
                # convert talk page result to wikidata(base)
                data = self.WD_convertContent(substed_content)
                #outpage = page.toggleTalkPage()
                outpage = pywikibot.wikidataPage(self.site, page.toggleTalkPage().title())
                #dic = json.loads(outpage.get())
                dic = outpage.getentities()

                # check for changes and then write/change/set values
                summary = u'Bot: update data because of configuration on %s.' % page.title(asLink=True)
                if not self.WD_save(outpage, dic[u'claims'], {u'p32': data}, summary):
                    pywikibot.output(u'NOTHING TO DO!')
            else:
                # if changed, write!
                if (substed_content != content):
                #if substed_tags:
                    self.outputContentDiff(content, substed_content)

                    head = i18n.twtranslate(self.site.lang,
                                            'thirdparty-drtrigonbot-sum_disc-summary-head')
                    if msg is None:
                        msg = i18n.twtranslate(self.site.lang,
                                               'thirdparty-drtrigonbot-subster-summary-mod')
                    flags = copy.deepcopy(EditFlags)
                    if page.title() in self._flagenable:
                        flags.update( self._flagenable[page.title()] )
                    pywikibot.output(u'Flags used for writing: %s' % flags)
                    self.save( page, substed_content,
                               (head + u' ' + msg) % {'tags':", ".join(substed_tags)},
                               **flags )
                else:
                    pywikibot.output(u'NOTHING TO DO!')

    def subContent(self, content, params):
        """Substitute the tags in content according to params.

           @param content: Content with tags to substitute.
           @type  content: string
           @param params: Params with data how to substitute tags.
           @type  params: dict

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_content = content
        substed_tags = []  # DRTRIGON-73

        # 0.) subst (internal) magic words
        try:
            (substed_content, tags) = self.subBotMagicWords(substed_content)
            substed_tags += tags
        except:
            exc_info = sys.exc_info()
            (exception_only, result) = pywikibot.gettraceback(exc_info)
            substed_content += ast.literal_eval(self._param_default['error']) %\
                               {'error': bot_config['ErrorTemplate'] %\
                                 ( pywikibot.Timestamp.now().isoformat(' '),
                                   u' ' + result.replace(u'\n', u'\n ').rstrip() ) }
            substed_tags.append( u'>error:BotMagicWords<' )

        if (len(params) == 1) and ast.literal_eval(params[0]['magicwords_only']):
            return (substed_content, substed_tags)

        for item in params:
            # 1.) - 5.) subst templates
            try:
                (substed_content, tags) = self.subTemplate(substed_content, item)
                substed_tags += tags
            except:
                exc_info = sys.exc_info()
                (exception_only, result) = pywikibot.gettraceback(exc_info)
                substed_content += ast.literal_eval(item['error']) %\
                                   {'error': bot_config['ErrorTemplate'] %\
                                     ( item['value'],
                                       pywikibot.Timestamp.now().isoformat(' '),
                                       u' ' + result.replace(u'\n', u'\n ').rstrip() ) }
                substed_tags.append( u'>error:%s<' % item['value'] )

        return (substed_content, substed_tags)

    def subBotMagicWords(self, content):
        """Substitute the DrTrigonBot Magic Word (tag)s in content.

           @param content: Content with tags to substitute.
           @type  content: string

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_tags = []  # DRTRIGON-73

        # 0.) subst (internal) magic words
        for subitem in magic_words.keys():
            prev_content = content
            content = self.get_var_regex(subitem).sub( (self._var_regex_str%{'var':subitem,'cont':magic_words[subitem]}),
                                                       content, 1)  # subst. once
            if (content != prev_content):
                substed_tags.append(subitem)

        return (content, substed_tags)

    def subTemplate(self, content, param):
        """Substitute the template tags in content according to param.

           @param content: Content with tags to substitute.
           @type  content: string
           @param param: Param with data how to substitute tags.
           @type  param: dict

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_tags = []  # DRTRIGON-73
        prev_content = content

        # 0.2.) check for 'simple' mode and get additional params
        if param['simple']:
            p = self.site.getExpandedString(param['simple'])
            param.update( pywikibot.extract_templates_and_params(p)[0][1] )

        # 0.5.) check cron/date
        if param['cron']:
            # [min] [hour] [day of month] [month] [day of week]
            # (date supported only, thus [min] and [hour] dropped)
            if not (param['cron'][0] == '@'):
                param['cron'] = '* * ' + param['cron']
            entry = crontab.CronTab(param['cron'])
            # find the delay from midnight (does not return 0.0 - but next)
            delay = entry.next(datetime.datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)-datetime.timedelta(microseconds=1))

            pywikibot.output(u'CRON delay for execution: %.3f (<= %i)' % (delay, bot_config['CRONMaxDelay']))

            if not (delay <= bot_config['CRONMaxDelay']):
                return (content, substed_tags)

        # 1.) getUrl or wiki text
        # (security: check url not to point to a local file on the server,
        #  e.g. 'file://' - same as used in xsalt.py)
        secure = False
        for item in [u'http://', u'https://', u'mail://']:
            secure = secure or (param['url'][:len(item)] == item)
        param['wiki'] = ast.literal_eval(param['wiki'])
        param['zip']  = ast.literal_eval(param['zip'])
        if (not secure) and (not param['wiki']):
            return (content, substed_tags)
        if   param['wiki']:
            if ast.literal_eval(param['expandtemplates']):  # DRTRIGON-93 (only with 'wiki')
                external_buffer = pywikibot.Page(self.site, param['url']).get(expandtemplates=True)
            else:
                external_buffer = self.load( pywikibot.Page(self.site, param['url']) )
        elif (param['url'][:7] == u'mail://'): # DRTRIGON-101
            param['url'] = param['url'].replace(u'{{@}}', u'@')     # e.g. nlwiki
            mbox = SubsterMailbox(pywikibot.config.datafilepath(bot_config['data_path'], bot_config['mbox_file'], ''))
            external_buffer = mbox.find_data(param['url'])
            mbox.close()
        elif param['zip']:
            external_buffer = urllib.urlopen(param['url']).read()
            # issue with r355: http://de.wikipedia.org/w/index.php?title=Vorlage:Infobox_Kreditinstitut/DatenDE&oldid=105472739
            #f_url, external_buffer = http.request(self.site, param['url'], no_hostname=True, back_response=True)
            #external_buffer = f_url.read()
            #del f_url   # free some memory (no need to keep a copy...)
        else:
            external_buffer = http.request(self.site, param['url'], no_hostname = True)

        # some intermediate processing (unzip, xlsx2csv, ...)
        if param['zip']:
            fileno          = 0 if (param['zip'] == True) else (param['zip']-1)
            external_buffer = self.unzip(external_buffer, fileno)
        if param['xlsx']:
            external_buffer = self.xlsx2csv(external_buffer, param['xlsx'])
        if param['ods']:
            external_buffer = self.ods2csv(external_buffer, param['ods'])

        if not ast.literal_eval(param['beautifulsoup']):    # DRTRIGON-88
            # 2.) regexp
            #for subitem in param['regex']:
            subitem = param['regex']
            regex = re.compile(subitem, re.S | re.I)

            # 3.) subst in content
            external_data = regex.search(external_buffer)

            external_data_dict = {}
            if external_data:    # not None
                external_data = external_data.groups()

                pywikibot.output(u'Groups found by regex: %i' % len(external_data))

                # DRTRIGON-114: Support for named groups in regexs
                if regex.groupindex:
                    for item in regex.groupindex:
                        external_data_dict[u'%s-%s' % (param['value'], item)] = external_data[regex.groupindex[item]-1]
                elif (len(external_data) == 1):
                    external_data_dict = {param['value']: external_data[0]}
                else:
                    external_data_dict = {param['value']: str(external_data)}
            logging.getLogger('subster').debug( str(external_data_dict) )

            param['postproc'] = eval(param['postproc'])
            # should be secured as given below, but needs code changes in wiki too
            #param['postproc'] = ast.literal_eval(param['postproc'])
            for value in external_data_dict:
                external_data = external_data_dict[value]

                # 4.) postprocessing
                func  = param['postproc'][0]    # needed by exec call of self._code
                DATA  = [ external_data ]       #
                args  = param['postproc'][1:]   #
                scope = {}                      # (scope to run in)
                scope.update( locals() )        # (add DATA, *args, ...)
                scope.update( globals() )       # (add imports and else)
                if func:
                    exec(self._code + (bot_config['CodeTemplate'] % func), scope, scope)
                    external_data = DATA[0]
                logging.getLogger('subster').debug( external_data )

                # 5.) subst content
                var_regex = self.get_var_regex(value)
                content = var_regex.sub((self._var_regex_str%{'var':value,'cont':external_data}), content, int(param['count']))
                if (content != prev_content):
                    substed_tags.append(value)
        else:
            # DRTRIGON-105: Support for multiple BS template configurations
            value = param['value']
            if value:
                value += u'-'

            # DRTRIGON-88: Enable Beautiful Soup power for Subster
            BS_tags = self.get_BS_regex(value).findall(content)

            pywikibot.output(u'BeautifulSoup tags found by regex: %i' % len(BS_tags))

            BS = BeautifulSoup.BeautifulSoup(external_buffer)
            for item in BS_tags:
                external_data = eval('BS.%s' % item[1])
                external_data = self._BS_regex_str%{'var1':value+'BS:'+item[1],'var2':value,'cont':external_data}
                content = content.replace(item[0], external_data, 1)

            if (content != prev_content):
                substed_tags.append(value+'BS')

        return (content, substed_tags)

    def outputContentDiff(self, content, substed_content):
        """Outputs the diff between the original and the new content.

           @param content: Original content.
           @type  content: string
           @param substed_content: New content.
           @type  substed_content: string

           Returns nothing, but outputs/prints the diff.
        """
        diff = difflib.Differ().compare(content.splitlines(1), substed_content.splitlines(1))
        diff = [ line for line in diff if line[0].strip() ]
        pywikibot.output(u'Diff:')
        pywikibot.output(u'--- ' * 15)
        pywikibot.output(u''.join(diff))
        pywikibot.output(u'--- ' * 15)

    def WD_convertContent(self, substed_content):
        """Converts the substed content to Wikidata format in order to save.
           (1 line of wiki text is converted to 1 claim/statement)

           @param substed_content: New content (with tags).
           @type  substed_content: string
        """
        # DRTRIGON-130: convert talk page result to wikidata(base)
        #res, i = {}, 0
        res = []
        for line in substed_content.splitlines():
            #data = self.get_var_regex('(.*?)', '(.*?)').findall(line)
            data = self.get_var_regex('.*?', '(.*?)').sub('\g<1>', line)
            #if not data:
            if data == line:
                continue
            #buf = []
            #for item in data:
            #    #print item[0], item[1]
            #    params = { u'property':  u'p%i' % i,
            #               u'value': item[1] }
            #    buf.append(params)
            #res[u'p%i' % i] = buf
            #i += 1
            res.append(data)

        return res

    def WD_save(self, outpage, dic, data, comment=None):
        """Stores the content to Wikidata.

           @param dic: Original content.
           @type  dic: dict
           @param data: New content.
           @type  data: dict

           Returns nothing, but stores the changed content.
        """
        # DRTRIGON-130: check for changes and then write/change/set values
        changed = False
        for prop in data:
            pywikibot.output(u'Checking claim with %i values' % len(data[prop]))
            for i, item in enumerate(data[prop]):
                if (i < len(dic[prop])) and \
                  (dic[prop][i][u'mainsnak'][u'datavalue'][u'value'] == item):
                    pass    # same value; nothing to do
                else:
                    # changes; update or create claim
                    changed = True
                    if (i < len(dic[prop])):
                        #print item, dic[prop][i][u'mainsnak'][u'datavalue'][u'value']
                        pywikibot.output(u'Updating claim with value: %s' % item)
                        outpage.setclaimvalue(dic[prop][i][u'id'], item, comment=comment)
                    else:
                        pywikibot.output(u'Creating new claim with value: %s' % item)
                        outpage.createclaim(prop, item, comment=comment)
        # speed-up by setting everything at once (in one single write attempt)
        #outpage.editentity(data = {u'claims': data})
        #outpage.setitem()

        return changed

    def get_var_regex(self, var, cont='.*?'):
        """Get regex used/needed to find the tags to replace.

           @param var: The tag/variable name.
           @type  var: string
           @param cont: The content/value of the variable.
           @type  cont: string

           Return the according (and compiled) regex object.
        """
        return re.compile((self._var_regex_str%{'var':var,'cont':cont}), re.S | re.I)

    def get_BS_regex(self, var, cont='(.*?)'):
        """Get regex used/needed to find the BS tags to replace.

           @param var: The tag/variable name.
           @type  var: string
           @param cont: The content/value of the variable.
           @type  cont: string

           Return the according (and compiled) regex object.
        """
        return re.compile(u'(' + self._BS_regex_str%{'var1':var+'BS:(.*?)','var2':var,'cont':cont} + u')')

    def unzip(self, external_buffer, i):
        """Convert zip data to plain format.
        """

        zip_buffer = zipfile.ZipFile(StringIO.StringIO(external_buffer))
        data_file  = zip_buffer.namelist()[i]
        external_buffer = zip_buffer.open(data_file).read().decode('latin-1')

        return external_buffer

    def xlsx2csv(self, external_buffer, sheet):
        """Convert xlsx (EXCEL) data to csv format.
        """

        wb = openpyxl.reader.excel.load_workbook(StringIO.StringIO(external_buffer), use_iterators = True)

        sheet_ranges = wb.get_sheet_by_name(name = sheet)

        output = StringIO.StringIO()
        spamWriter = csv.writer(output)

        for row in sheet_ranges.iter_rows(): # it brings a new method: iter_rows()
            spamWriter.writerow([ cell.internal_value for cell in row ])

        external_buffer = output.getvalue()
        output.close()

        return external_buffer

    def ods2csv(self, external_buffer, sheet):
        """Convert ods (Open/Libre Office) data to csv format.
        """
        # http://www.mail-archive.com/python-list@python.org/msg209447.html

        import odf
        from odf import opendocument, table, teletype

        doc = odf.opendocument.load(StringIO.StringIO(external_buffer))

        output = StringIO.StringIO()
        spamWriter = csv.writer(output)

        for sheet in doc.getElementsByType(odf.table.Table):
            if not (sheet.getAttribute('name') == sheet):
                continue
            for row in sheet.getElementsByType(odf.table.TableRow):
                spamWriter.writerow([ odf.teletype.extractText(cell).encode('utf-8')
                                      for cell in row.getElementsByType(odf.table.TableCell) ])

        external_buffer = output.getvalue()
        output.close()

        return external_buffer


class SubsterMailbox(mailbox.mbox):
    def __init__(self, mbox_file):
        mailbox.mbox.__init__(self, mbox_file)
        self.lock()

        self.remove_duplicates()

    def remove_duplicates(self):
        """Find mails with same 'From' (sender) and remove all
        except the most recent one.
        """

        unique = {}
        remove = []
        for i, message in enumerate(self):
            sender   = message['from']       # Could possibly be None.
            timestmp = message['date']       # Could possibly be None.

            timestmp = time.mktime( email.utils.parsedate(timestmp) )
            timestmp = datetime.datetime.fromtimestamp( timestmp )

            if sender in unique:
                (j, timestmp_j) = unique[sender]

                if (timestmp >= timestmp_j):
                    remove.append( j )
                else:
                    remove.append( i )
            else:
                unique[sender] = (i, timestmp)

        remove.reverse()
        for i in remove:
            self.remove(i)

        self.flush()
        #self.close()

        if remove:
            pywikibot.output('Removed %i depreciated email data source(s).' % len(remove))

    def find_data(self, url):
        """Find mail according to given 'From' (sender).
        """

        url = (url[:7], ) + tuple(url[7:].split('/'))
        content = []

        for i, message in enumerate(self):
            sender   = message['from']          # Could possibly be None.
            subject  = message['subject']       # Could possibly be None.
            timestmp = message['date']       # Could possibly be None.

            if sender and url[1] in sender:
                # data found
                pywikibot.output('Found email data source:')
                pywikibot.output('%i / %s / %s / %s' % (i, sender, subject, timestmp))

                full = (url[2] == 'attachment-full')
                ind  = 0    # default; ignore attachement index
                if   (url[2] == 'all'):
                    content = [ message.as_string(True) ]
                elif (url[2] == 'attachment') or full:
                    if len(url) > 3:
                        ind = int(url[3])   # combine 'ind' with 'full=True'...?
                    counter = 1
                    content = []
                    for part in message.walk():
                        # multipart/* are just containers
                        if part.get_content_maintype() == 'multipart':
                            continue
                        # Applications should really sanitize the given filename so that an
                        # email message can't be used to overwrite important files
                        filename = part.get_filename()
                        if filename or full:
                            if not filename:
                                ext = mimetypes.guess_extension(part.get_content_type())
                                if not ext:
                                    # Use a generic bag-of-bits extension
                                    ext = '.bin'
                                filename = 'part-%03d%s' % (counter, ext)

                            content += [ part.get_payload(decode=True) ]
                            pywikibot.output('Found attachment # %i: "%s"' % (counter, filename))

                            if counter == ind:
                                return content[-1]

                            counter += 1

                            if (not full) and (not ind):
                                break

                    break

        return string.join(content)


def main():
    args = pywikibot.handleArgs()
    bot  = SubsterBot()   # for several user's, but what about complete automation (continous running...)
    if len(args) > 0:
        for arg in args:
            pywikibot.showHelp()
            return
    try:
        bot.run()
    except KeyboardInterrupt:
        pywikibot.output('\nQuitting program...')

if __name__ == "__main__":
    try:
        main()
    finally:
        pywikibot.stopme()

